"""Financial Spreadsheet Stock Price Updater

This script automates the maintenance of a stock spreadsheet.
The script reads in a spreadsheet file (*.xlsx), with each row a stock holding.
The second column is the ticker symbol. The third column is the stock price.

The script uses the yahoo finance api to find the current stock price.
It updates column three on each stock row with that current price.

The updated spreadsheet is saved in the same directory.
The new file name has the current date to the file name.

Functions:
    * update_spreadsheet(fname) - updates spreadsheet. returns null
    * get_price(ticker) - takes in ticker symbol, returns current price for ticker
    * format_outputfile_name(fname) - creates new output file name
    * main - the main function of the script
"""
import yfinance as yf
import math
import sys
import openpyxl as xl
import re
import time
import shutil

def main(argv):
    """ reads in stock spreadsheet and calls functions to update it.
        Take stock spreadhseet path in as a parameter.
        Call function to collect new price data and update spreadsheet.
        create new saved filename with current date appended

        Functions:
        * update_spreadsheet(fname) - updates spreadsheet. returns null

    """
    print("Processing file:", argv[0])

    fname = argv[0]
    update_spreadsheet(fname)
    return

def format_outputfile_name(fname):
    """ create new saved filename with current date appended"""
    new_date = time.strftime("%Y%m%d")
    sep = '.'
    stub = fname.split(sep, 1)[0]
    npath = stub + '_' + new_date +'.xlsx'
    print("New File : ", npath)

    return npath

def get_price(ticker):
    """ Gets the latest price for a stock

        Takes in ticker symbol and returns latest stock price from yahoo finanace API
    """
    price = float('nan')
    x = 0
    data = yf.download(tickers=ticker, period='1wk', interval='1d')
    while math.isnan(price):
        x = x - 1
        price = data['Close'][x]
    print('Price : ', price)
    return price

def update_spreadsheet(fname):
    """ Updates stock price for all ticker symbols in financial spreadsheet.
        update_spreadsheet(fname):
        Reads in original spreadsheet file, where each row is a stock holding.
        Calls get_price() function to get current price for ticker (column 2).
        Updates the stock price in spreadsheet (column 4).
        Saves new spreadsheet by appending current date on filename.
        Returns null.

        Functions:
        * get_price(ticker) - takes in ticker symbol, returns current price for ticker
        * format_outputfile_name(fname) - creates new output file name

    """
    path = fname
    wb_obj = xl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    min_column = sheet_obj.min_column
    print('Max : ', max_column, ' Min : ', min_column, ' Rows : ', max_row )
    for j in range(2,max_row):
        if sheet_obj.cell(row = j, column=2).value:
            ticker = sheet_obj.cell(row = j, column=2).value
            price = sheet_obj.cell(row=j, column=4).value
            new_price = get_price(ticker)
            print('Stock : ', ticker, 'Price : ', price, 'New Price : ', new_price,)
#           print('Column Value: ', sheet_obj.cell(row = j, column=2).value)
            price_cell = sheet_obj.cell(row=j, column=4)
            price_cell.value = new_price
    new_fname = format_outputfile_name(fname)
    wb_obj.save(new_fname)
    return

if __name__ == '__main__':
    main(sys.argv[1:])
